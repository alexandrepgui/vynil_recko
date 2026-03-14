"""Collection export generators: CSV, Excel (.xlsx), and PDF."""

from __future__ import annotations

import base64
import csv
import io
import logging
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from html import escape
from pathlib import Path

import requests
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font

from config import DISCOGS_USER_AGENT
from repository.models import CollectionItem

log = logging.getLogger(__name__)

EXPORT_COLUMNS = ["Artist", "Title", "Year", "Format", "Genres", "Styles", "Date Added", "Discogs URL"]

DISCOGS_URL_TEMPLATE = "https://www.discogs.com/release/{}"


def _discogs_url(release_id: int) -> str:
    return DISCOGS_URL_TEMPLATE.format(release_id)


def _format_list(items: list[str]) -> str:
    return ", ".join(items)


def _format_date(date_str: str | None) -> str:
    if not date_str:
        return ""
    try:
        dt = datetime.fromisoformat(date_str)
        return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return date_str


def _item_row(item: CollectionItem) -> list[str]:
    return [
        item.artist,
        item.title,
        str(item.year) if item.year else "",
        item.format,
        _format_list(item.genres),
        _format_list(item.styles),
        _format_date(item.date_added),
        _discogs_url(item.release_id),
    ]


# ── CSV ──────────────────────────────────────────────────────────────────────


def generate_csv(items: list[CollectionItem]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(EXPORT_COLUMNS)
    for item in items:
        writer.writerow(_item_row(item))
    # UTF-8 BOM so Excel opens with correct encoding
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


# ── Excel ────────────────────────────────────────────────────────────────────


def generate_xlsx(items: list[CollectionItem]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Collection"

    # Header row
    ws.append(EXPORT_COLUMNS)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    # Data rows
    for item in items:
        row = _item_row(item)
        ws.append(row)
        # Make Discogs URL a clickable hyperlink
        url_cell = ws.cell(row=ws.max_row, column=len(EXPORT_COLUMNS))
        url_cell.hyperlink = row[-1]
        url_cell.font = Font(color="0563C1", underline="single")

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF (WeasyPrint — HTML/CSS to PDF) ───────────────────────────────────────

_MAX_IMG_WORKERS = 4
_COVER_SIZE = 140  # 2x the 70px display size for sharpness
_FONTS_DIR = Path(__file__).resolve().parent.parent / "assets" / "fonts"
_ICON_SVG = Path(__file__).resolve().parent.parent.parent / "frontend" / "src" / "assets" / "icon.svg"


def _download_cover_b64(url: str) -> str | None:
    """Download a cover image, resize to thumbnail, and return as base64 data URI."""
    try:
        resp = requests.get(url, timeout=10, headers={"User-Agent": DISCOGS_USER_AGENT})
        resp.raise_for_status()
        img = Image.open(io.BytesIO(resp.content))
        img.thumbnail((_COVER_SIZE, _COVER_SIZE))
        img = img.convert("RGB")
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=70, optimize=True)
        b64 = base64.b64encode(buf.getvalue()).decode()
        return f"data:image/jpeg;base64,{b64}"
    except Exception:
        log.warning("Failed to download cover from %s", url, exc_info=True)
        return None


def _download_all_covers(items: list[CollectionItem]) -> dict[int, str]:
    """Download cover images in parallel. Returns mapping of instance_id → data URI."""
    covers: dict[int, str] = {}
    futures = {}
    with ThreadPoolExecutor(max_workers=_MAX_IMG_WORKERS) as pool:
        for item in items:
            url = item.custom_cover_image or item.cover_image
            if not url:
                continue
            futures[pool.submit(_download_cover_b64, url)] = item.instance_id
        for future in as_completed(futures):
            instance_id = futures[future]
            result = future.result()
            if result:
                covers[instance_id] = result
    return covers


def _group_by_format(items: list[CollectionItem]) -> list[tuple[str, list[CollectionItem]]]:
    """Group items by format, sorted by group name. Items within each group sorted by artist."""
    groups: dict[str, list[CollectionItem]] = {}
    for item in items:
        key = item.format or "Other"
        groups.setdefault(key, []).append(item)
    for group_items in groups.values():
        group_items.sort(key=lambda i: (i.artist.lower(), i.year))
    return sorted(groups.items(), key=lambda g: g[0].lower())


def _read_icon_svg() -> str:
    """Read the app icon SVG and return as inline data URI."""
    try:
        svg_bytes = _ICON_SVG.read_bytes()
        b64 = base64.b64encode(svg_bytes).decode()
        return f"data:image/svg+xml;base64,{b64}"
    except Exception:
        log.warning("Failed to read icon SVG from %s", _ICON_SVG, exc_info=True)
        return ""


_PDF_CSS = """
@font-face {
    font-family: 'Shrikhand';
    src: url('file://FONTS_DIR/Shrikhand-Regular.ttf');
}
@font-face {
    font-family: 'DMSans';
    src: url('file://FONTS_DIR/DMSans.ttf');
}
@font-face {
    font-family: 'JBMono';
    src: url('file://FONTS_DIR/JetBrainsMono.ttf');
}

@page {
    size: A4;
    margin: 12mm 15mm 15mm 15mm;
    background: #1A1816;

    @bottom-center {
        font-family: 'JBMono', monospace;
        font-size: 7pt;
        color: #787878;
        content: counter(page) " / " counter(pages);
    }
}

@page :first {
    @bottom-center { content: none; }
}

* { box-sizing: border-box; margin: 0; padding: 0; }

body {
    font-family: 'DMSans', sans-serif;
    color: #FFFFFF;
    background: #1A1816;
}

/* ── Title page ──────────────────────────────── */

.title-page {
    page-break-after: always;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    height: 100%;
    text-align: center;
}

.title-icon {
    width: 48px;
    height: 48px;
    margin-bottom: 16px;
}

.title-wordmark {
    font-family: 'Shrikhand', cursive;
    font-size: 32pt;
    font-weight: 400;
    letter-spacing: 0.01em;
    margin-bottom: 6px;
}

.title-user {
    font-size: 13pt;
    color: #B4B4B4;
    margin-bottom: 20px;
}

.title-stats {
    font-family: 'JBMono', monospace;
    font-size: 8.5pt;
    color: #787878;
    margin-bottom: 14px;
}

.title-divider {
    width: 50px;
    height: 1px;
    background: #3D3831;
    margin-bottom: 0;
}

.title-footer {
    position: absolute;
    bottom: 0;
    left: 0;
    right: 0;
    text-align: center;
    padding-bottom: 5mm;
}

.title-footer a {
    font-size: 8pt;
    color: #787878;
    text-decoration: none;
}

/* ── Page header (pages 2+) ──────────────────── */

.page-header {
    font-family: 'DMSans', sans-serif;
    font-size: 7pt;
    color: #787878;
    text-align: right;
    margin-bottom: 8px;
}

/* ── Group header ────────────────────────────── */

.group {
    margin-bottom: 4px;
}

.group-header {
    display: flex;
    align-items: baseline;
    gap: 10px;
    border-top: 1px solid #3D3831;
    padding-top: 8px;
    margin-top: 10px;
    margin-bottom: 6px;
}

.group-name {
    font-size: 12pt;
    font-weight: 500;
    color: #FFFFFF;
}

.group-count {
    font-family: 'JBMono', monospace;
    font-size: 7pt;
    color: #787878;
}

/* ── Record card ─────────────────────────────── */

.card {
    display: flex;
    gap: 10px;
    padding: 6px;
    background: #262320;
    border-radius: 4px;
    margin-bottom: 4px;
    page-break-inside: avoid;
}

.card-cover {
    width: 70px;
    height: 70px;
    object-fit: cover;
    border-radius: 2px;
    flex-shrink: 0;
}

.card-cover-placeholder {
    width: 70px;
    height: 70px;
    background: #32302B;
    border-radius: 2px;
    flex-shrink: 0;
}

.card-info {
    flex: 1;
    min-width: 0;
    display: flex;
    flex-direction: column;
    justify-content: center;
    gap: 1px;
}

.card-title {
    font-size: 9.5pt;
    font-weight: 500;
    color: #FFFFFF;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}

.card-artist {
    font-size: 8.5pt;
    color: #B4B4B4;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
    margin-bottom: 3px;
}

.card-tags {
    display: flex;
    flex-wrap: wrap;
    gap: 3px;
    align-items: flex-start;
}

.tag {
    font-family: 'JBMono', monospace;
    font-size: 6pt;
    line-height: 6pt;
    height: 10pt;
    display: inline-flex;
    align-items: center;
    padding: 0 5px;
    background: #32302B;
    color: #787878;
    border-radius: 2px;
    white-space: nowrap;
}

/* ── Title page stats ────────────────────────── */

.title-stats-grid {
    display: flex;
    justify-content: center;
    gap: 20px;
    margin-top: 24px;
    margin-bottom: 20px;
}

.stat-card {
    background: #262320;
    border-radius: 6px;
    padding: 12px 18px;
    text-align: center;
    min-width: 80px;
}

.stat-value {
    font-family: 'JBMono', monospace;
    font-size: 16pt;
    color: #FFFFFF;
    margin-bottom: 2px;
}

.stat-label {
    font-size: 7pt;
    color: #787878;
    text-transform: uppercase;
    letter-spacing: 0.05em;
}

.title-breakdown {
    margin-top: 8px;
    margin-bottom: 16px;
}

.title-breakdown-row {
    display: flex;
    justify-content: center;
    gap: 12px;
    margin-bottom: 4px;
}

.breakdown-item {
    font-family: 'JBMono', monospace;
    font-size: 7.5pt;
    color: #787878;
}

.breakdown-item strong {
    color: #B4B4B4;
    font-weight: 500;
}
""".replace("FONTS_DIR", str(_FONTS_DIR))


def _compute_stats(items: list[CollectionItem], groups: list[tuple[str, list[CollectionItem]]]) -> dict:
    """Compute descriptive statistics for the collection."""
    artists = set()
    genre_counts: dict[str, int] = {}
    decades: dict[str, int] = {}
    for item in items:
        artists.add(item.artist)
        for g in item.genres:
            genre_counts[g] = genre_counts.get(g, 0) + 1
        if item.year:
            decade = f"{(item.year // 10) * 10}s"
            decades[decade] = decades.get(decade, 0) + 1

    top_genres = sorted(genre_counts.items(), key=lambda x: -x[1])[:5]
    top_decades = sorted(decades.items(), key=lambda x: -x[1])[:4]

    return {
        "artist_count": len(artists),
        "format_breakdown": [(name, len(g_items)) for name, g_items in groups],
        "top_genres": top_genres,
        "top_decades": top_decades,
    }


def _breakdown_html(items: list[tuple[str, int]]) -> str:
    """Render a list of (label, count) pairs as middot-separated breakdown spans."""
    return "  &middot;  ".join(
        f'<span class="breakdown-item"><strong>{count}</strong> {escape(label)}</span>'
        for label, count in items
    )


def generate_pdf(items: list[CollectionItem], username: str = "collector") -> bytes:
    from weasyprint import HTML

    covers = _download_all_covers(items)
    groups = _group_by_format(items)
    icon_uri = _read_icon_svg()
    now = datetime.now(timezone.utc).strftime("%B %d, %Y")
    stats = _compute_stats(items, groups)

    parts: list[str] = []

    format_html = _breakdown_html(stats["format_breakdown"])
    genre_html = _breakdown_html(stats["top_genres"])
    decade_html = _breakdown_html(stats["top_decades"])

    # Title page
    parts.append(f"""
    <div class="title-page">
        <img class="title-icon" src="{icon_uri}" alt="">
        <div class="title-wordmark">groove log</div>
        <div class="title-user">{escape(username)}\u2019s collection</div>

        <div class="title-stats-grid">
            <div class="stat-card">
                <div class="stat-value">{len(items)}</div>
                <div class="stat-label">Records</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{stats["artist_count"]}</div>
                <div class="stat-label">Artists</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{len(stats["format_breakdown"])}</div>
                <div class="stat-label">Formats</div>
            </div>
        </div>

        <div class="title-breakdown">
            <div class="title-breakdown-row">{format_html}</div>
        </div>

        <div class="title-divider"></div>

        <div class="title-breakdown" style="margin-top: 16px;">
            <div class="title-breakdown-row">{genre_html}</div>
        </div>
        <div class="title-breakdown">
            <div class="title-breakdown-row">{decade_html}</div>
        </div>

        <div class="title-stats">{now}</div>

        <div class="title-footer">
            <a href="https://groovelog.app">made with groove log</a>
        </div>
    </div>
    """)

    # Groups
    for group_name, group_items in groups:
        parts.append(f"""
        <div class="group">
            <div class="group-header">
                <span class="group-name">{escape(group_name)}</span>
                <span class="group-count">{len(group_items)} record{"s" if len(group_items) != 1 else ""}</span>
            </div>
        """)
        for item in group_items:
            cover_uri = covers.get(item.instance_id, "")
            tags = []
            if item.year:
                tags.append(str(item.year))
            tags.extend(item.genres[:3])
            tags.extend(item.styles[:3])
            tags_html = "".join(f'<span class="tag">{escape(t)}</span>' for t in tags)

            if cover_uri:
                cover_html = f'<img class="card-cover" src="{cover_uri}" alt="">'
            else:
                cover_html = '<div class="card-cover-placeholder"></div>'

            parts.append(f"""
            <div class="card">
                {cover_html}
                <div class="card-info">
                    <div class="card-title">{escape(item.title)}</div>
                    <div class="card-artist">{escape(item.artist)}</div>
                    <div class="card-tags">{tags_html}</div>
                </div>
            </div>
            """)
        parts.append("</div>")

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><style>{_PDF_CSS}</style></head>
<body>{"".join(parts)}</body></html>"""

    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=True) as tmp:
        HTML(string=html).write_pdf(target=tmp.name)
        return Path(tmp.name).read_bytes()
