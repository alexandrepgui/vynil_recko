"""Unit tests for services/export.py — CSV, Excel, PDF generation."""

import csv
import io
import zipfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import load_workbook

from repository.models import CollectionItem
from services.export import (
    generate_csv,
    generate_pdf,
    generate_xlsx,
)


def _make_item(**overrides) -> CollectionItem:
    defaults = dict(
        user_id="u1",
        instance_id=1,
        release_id=100,
        title="Test Album",
        artist="Test Artist",
        year=2020,
        genres=["Rock"],
        styles=["Indie"],
        format="LP",
        cover_image="https://img.example.com/cover.jpg",
        date_added="2024-06-01T12:00:00+00:00",
    )
    defaults.update(overrides)
    return CollectionItem(**defaults)


ITEMS = [
    _make_item(instance_id=1, release_id=100, artist="Miles Davis", title="Kind of Blue", year=1959, genres=["Jazz"], styles=["Modal"]),
    _make_item(instance_id=2, release_id=200, artist="Radiohead", title="OK Computer", year=1997, genres=["Electronic", "Rock"], styles=["Alternative Rock"]),
    _make_item(instance_id=3, release_id=300, artist="Björk", title="Homogenic", year=1997, genres=["Electronic"], styles=["IDM", "Trip Hop"], date_added=None),
]


# ── CSV ──────────────────────────────────────────────────────────────────────


class TestGenerateCSV:

    def test_basic_output(self):
        data = generate_csv(ITEMS)
        # Starts with UTF-8 BOM
        assert data[:3] == b"\xef\xbb\xbf"

        text = data.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)

        assert rows[0] == ["Artist", "Title", "Year", "Format", "Genres", "Styles", "Date Added", "Discogs URL"]
        assert len(rows) == 4

    def test_special_characters(self):
        items = [_make_item(artist="Björk", title="Jóga")]
        data = generate_csv(items)
        text = data.decode("utf-8-sig")
        assert "Björk" in text
        assert "Jóga" in text

    def test_empty(self):
        data = generate_csv([])
        text = data.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        assert len(rows) == 1  # header only

    def test_multi_genre(self):
        data = generate_csv(ITEMS)
        text = data.decode("utf-8-sig")
        assert "Electronic, Rock" in text

    def test_missing_date(self):
        data = generate_csv(ITEMS)
        text = data.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        # Björk has no date_added
        assert rows[3][6] == ""

    def test_discogs_url_present(self):
        data = generate_csv(ITEMS)
        text = data.decode("utf-8-sig")
        assert "https://www.discogs.com/release/100" in text


# ── Excel ────────────────────────────────────────────────────────────────────


class TestGenerateXLSX:

    def test_basic_output(self):
        data = generate_xlsx(ITEMS)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active

        assert ws.title == "Collection"
        assert ws.max_row == 4
        assert ws.cell(1, 1).value == "Artist"
        assert ws.cell(2, 1).value == "Miles Davis"

    def test_header_bold(self):
        data = generate_xlsx(ITEMS)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.cell(1, 1).font.bold is True

    def test_freeze_panes(self):
        data = generate_xlsx(ITEMS)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.freeze_panes == "A2"

    def test_hyperlink(self):
        data = generate_xlsx(ITEMS)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        url_cell = ws.cell(2, 8)
        assert url_cell.hyperlink is not None

    def test_empty(self):
        data = generate_xlsx([])
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.max_row == 1

    def test_special_characters(self):
        items = [_make_item(artist="Björk")]
        data = generate_xlsx(items)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.cell(2, 1).value == "Björk"


# ── PDF ──────────────────────────────────────────────────────────────────────


class TestGeneratePDF:

    def test_valid_pdf(self):
        data = generate_pdf(ITEMS)
        assert data[:5] == b"%PDF-"
        assert len(data) > 100

    def test_empty_collection(self):
        data = generate_pdf([])
        assert data[:5] == b"%PDF-"

    def test_items_without_cover(self):
        items = [_make_item(cover_image=None, custom_cover_image=None)]
        data = generate_pdf(items)
        assert data[:5] == b"%PDF-"

    @patch("services.export.requests.get")
    def test_cover_download_failure_handled(self, mock_get):
        """PDF generation shouldn't fail if cover downloads fail."""
        mock_get.side_effect = Exception("Network error")
        items = [_make_item(cover_image="https://img.example.com/fail.jpg")]
        data = generate_pdf(items)
        assert data[:5] == b"%PDF-"

    @patch("services.export.requests.get")
    def test_cover_download_success(self, mock_get):
        """PDF should attempt to download covers."""
        # Create a minimal valid PNG (1x1 pixel)
        png_data = (
            b"\x89PNG\r\n\x1a\n"  # PNG signature
            b"\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde"
            b"\x00\x00\x00\x0cIDATx\x9cc\xf8\x0f\x00\x00\x01\x01\x00\x05\x18\xd8N"
            b"\x00\x00\x00\x00IEND\xaeB`\x82"
        )
        mock_resp = MagicMock()
        mock_resp.content = png_data
        mock_resp.raise_for_status = MagicMock()
        mock_get.return_value = mock_resp

        items = [_make_item(cover_image="https://img.example.com/cover.png")]
        data = generate_pdf(items)
        assert data[:5] == b"%PDF-"
        mock_get.assert_called_once()
