"""Tests for GET /api/collection/export — CSV, Excel, PDF."""

import csv
import io
from unittest.mock import MagicMock

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from conftest import TEST_USER_ID
from repository.models import CollectionItem


SAMPLE_ITEMS = [
    CollectionItem(
        user_id=TEST_USER_ID,
        instance_id=100,
        release_id=555,
        title="Kind of Blue",
        artist="Miles Davis",
        year=1959,
        genres=["Jazz"],
        styles=["Modal"],
        format="LP",
        cover_image="https://img.discogs.com/cover1.jpg",
        date_added="2024-01-15T10:30:00+00:00",
    ),
    CollectionItem(
        user_id=TEST_USER_ID,
        instance_id=200,
        release_id=666,
        title="OK Computer",
        artist="Radiohead",
        year=1997,
        genres=["Electronic", "Rock"],
        styles=["Alternative Rock"],
        format="LP",
        cover_image="https://img.discogs.com/cover2.jpg",
        date_added="2024-02-20T14:00:00+00:00",
    ),
    CollectionItem(
        user_id=TEST_USER_ID,
        instance_id=300,
        release_id=777,
        title="Daft Punk",
        artist="Homework",
        year=1997,
        genres=["Electronic"],
        styles=["House"],
        format="CD",
        date_added=None,
    ),
]


@pytest.fixture()
def mock_repo():
    repo = MagicMock()
    repo.find_collection_items.return_value = list(SAMPLE_ITEMS)
    return repo


@pytest.fixture()
def client(mock_repo, mock_jwt_user):
    from deps import get_repo
    from main import app

    app.dependency_overrides[get_repo] = lambda: mock_repo
    yield TestClient(app)
    app.dependency_overrides.pop(get_repo, None)


# ── CSV ──────────────────────────────────────────────────────────────────────


def test_export_csv(client, mock_repo):
    resp = client.get("/api/collection/export?format=csv")
    assert resp.status_code == 200
    assert "text/csv" in resp.headers["content-type"]
    assert 'filename="groove-log-collection.csv"' in resp.headers["content-disposition"]

    # Parse CSV (skip BOM)
    text = resp.content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)

    assert rows[0] == ["Artist", "Title", "Year", "Format", "Genres", "Styles", "Date Added", "Discogs URL"]
    assert len(rows) == 4  # header + 3 items
    assert rows[1][0] == "Miles Davis"
    assert rows[1][1] == "Kind of Blue"
    assert rows[1][6] == "2024-01-15"
    assert rows[2][4] == "Electronic, Rock"


def test_export_csv_empty(client, mock_repo):
    mock_repo.find_collection_items.return_value = []
    resp = client.get("/api/collection/export?format=csv")
    assert resp.status_code == 200

    text = resp.content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    assert len(rows) == 1  # header only


def test_export_csv_passes_filters(client, mock_repo):
    client.get("/api/collection/export?format=csv&sort=title&sort_order=desc&q=jazz")
    call_kwargs = mock_repo.find_collection_items.call_args
    assert call_kwargs.kwargs["sort"] == "title"
    assert call_kwargs.kwargs["sort_order"] == "desc"
    assert call_kwargs.kwargs["query"] == "jazz"
    assert call_kwargs.kwargs["skip"] == 0
    assert call_kwargs.kwargs["limit"] == 0


# ── Excel ────────────────────────────────────────────────────────────────────


def test_export_xlsx(client, mock_repo):
    resp = client.get("/api/collection/export?format=xlsx")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert 'filename="groove-log-collection.xlsx"' in resp.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.title == "Collection"

    # Header row
    headers = [cell.value for cell in ws[1]]
    assert headers == ["Artist", "Title", "Year", "Format", "Genres", "Styles", "Date Added", "Discogs URL"]

    # Data rows
    assert ws.max_row == 4  # header + 3 items
    assert ws.cell(2, 1).value == "Miles Davis"
    assert ws.cell(2, 2).value == "Kind of Blue"

    # Discogs URL should be a hyperlink
    url_cell = ws.cell(2, 8)
    assert url_cell.hyperlink is not None
    assert "555" in url_cell.hyperlink.target


def test_export_xlsx_empty(client, mock_repo):
    mock_repo.find_collection_items.return_value = []
    resp = client.get("/api/collection/export?format=xlsx")
    assert resp.status_code == 200

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row == 1  # header only


# ── PDF ──────────────────────────────────────────────────────────────────────


def test_export_pdf(client, mock_repo):
    # PDF with no cover image downloads (items have URLs but we mock requests)
    mock_repo.find_collection_items.return_value = [
        CollectionItem(
            user_id=TEST_USER_ID,
            instance_id=100,
            release_id=555,
            title="Kind of Blue",
            artist="Miles Davis",
            year=1959,
            genres=["Jazz"],
            format="LP",
        ),
    ]
    resp = client.get("/api/collection/export?format=pdf")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert 'filename="groove-log-collection.pdf"' in resp.headers["content-disposition"]
    # Valid PDF
    assert resp.content[:5] == b"%PDF-"


def test_export_pdf_empty(client, mock_repo):
    mock_repo.find_collection_items.return_value = []
    resp = client.get("/api/collection/export?format=pdf")
    assert resp.status_code == 200
    assert resp.content[:5] == b"%PDF-"


# ── Validation ───────────────────────────────────────────────────────────────


def test_export_invalid_format(client):
    resp = client.get("/api/collection/export?format=xml")
    assert resp.status_code == 422


def test_export_missing_format(client):
    resp = client.get("/api/collection/export")
    assert resp.status_code == 422


def test_export_requires_auth(mock_repo):
    """Verify that the endpoint requires authentication."""
    from deps import get_repo
    from main import app

    app.dependency_overrides[get_repo] = lambda: mock_repo
    unauthenticated_client = TestClient(app)

    resp = unauthenticated_client.get("/api/collection/export?format=csv")
    assert resp.status_code in (401, 403)

    app.dependency_overrides.pop(get_repo, None)
